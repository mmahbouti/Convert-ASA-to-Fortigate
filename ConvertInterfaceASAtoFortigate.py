import getpass
import colorama
from netmiko import ConnectHandler
from colorama import init
init()
from colorama import Fore,Back,Style
from openpyxl import load_workbook
from xlsxwriter import Workbook
import re
import requests
import urllib3
import json

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def name_evaluator(interface_name):
    try:
        if interface_name == 'no_interface_name':
            return True
        elif len(interface_name)> 15:
            return False
        elif interface_name.find("<") == 0 or interface_name.find(">") == 0 or interface_name.find("(") == 0 or interface_name.find(")") == 0 or interface_name.find("#") == 0 or interface_name.find("'") == 0 or interface_name.find('"') == 0 or interface_name.find(":") == 0 or interface_name.find("%") == 0:
            return False
        else:
            return True
    except Exception as e:
        print("name_evaluator failed!")
        print(str(e))


def find_type_interface(interface_id):
    try:
        type_interface = None

        portchannel_subinterface_pattern = re.compile(r'Port-channel\d+.\d+')
        portchannel_interface_pattern = re.compile(r'Port-channel\d+')

        if interface_id.startswith("GigabitEthernet") or interface_id.startswith("TenGigabitEthernet") or interface_id.startswith("FastEthernet"):
            type_interface = "physical_type"
        if interface_id.startswith("Management"):
            type_interface = "management"
        if bool(portchannel_subinterface_pattern.findall(interface_id)):
            type_interface = "portchannel_subinterface_type"
        elif bool(portchannel_interface_pattern.findall(interface_id)):
            type_interface = "portchannel_interface_type"

        return type_interface

    except Exception as e:
        print("find_type_interface failed!")
        print(str(e))


def edit_dict_list(input_dict_list):
    try:
        channel_group_dict = {}
        output_dict_list = []
        for entity in input_dict_list:
            interface_id = entity["interface_id"]
            channel_group = entity["channel-group"]
            interface_name = entity["interface_name"]

            if find_type_interface(interface_id) == "management":
                interface_id_fortigate = input(f"Enter corresponding interface physical {interface_id} in your fortigate: ")
                entity["interface_id"] = interface_id_fortigate
                entity["interface_name"] = interface_id_fortigate
                entity["interface_allow_access"] = 'ping https ssh http snmp'
                entity["interface"] = None
                entity["role"] = 'lan'
                entity["type"] = 'physical'
            elif find_type_interface(interface_id) == "physical_type":
                interface_id_fortigate = input(f"Enter corresponding interface physical {interface_id} in your fortigate: ")
                entity["interface_id"] = interface_id_fortigate
                entity["interface_name"] = interface_id_fortigate
                entity["interface_allow_access"] = 'ping'
                entity["interface"] = None
                entity["role"] = 'lan'
                entity["type"] = 'physical'
                entity["channel-group"] = None
                tmp_key = "Port-channel"+channel_group
                
                if tmp_key in channel_group_dict:
                    channel_group_dict[tmp_key] = channel_group_dict[tmp_key] + [interface_id_fortigate]

                else:
                    channel_group_dict[tmp_key] = [interface_id_fortigate]
                
            elif find_type_interface(interface_id) == "portchannel_interface_type":
                entity["interface"] = None
                entity["interface_allow_access"] = 'ping'
                entity["role"] = 'lan'
                entity["type"] = 'aggregate'
                entity["channel-group"] = json.dumps(channel_group_dict[interface_id])
                
            elif find_type_interface(interface_id) == "portchannel_subinterface_type":
                entity["role"] = 'lan'
                entity["interface_allow_access"] = 'ping'
                entity["type"] = 'vlan'
                entity["interface"] = interface_id.split(".")[0]

                
                
            if name_evaluator(interface_name)== False:
                interface_name_fortigate = input(f"Enter corresponding interface name {interface_name} in your fortigate: ")
                entity["interface_name_asa"] = interface_name
                entity["interface_name"] = interface_name_fortigate
            else:
                entity["interface_name_asa"] = interface_name
                
            output_dict_list.append(entity)
        #print(channel_group_dict)
        return output_dict_list
    except Exception as e:
        print("edit_dict_list failed!")
        print(str(e))


def test_interface(in_text):
    try:
        in_text = in_text[1:]
        input_txt_list = in_text.split("\n")
        output_txt_list = []
        temp = {}
        for line in input_txt_list:
            d = {}

            if line.startswith("interface"):
                interface_id = line.split(" ")[1]
                d.update({"interface_id": interface_id})
                temp.update(d)

            elif line.startswith(" vlan"):
                vlan_id = line.split(" ")[2]
                d.update({"vlan": vlan_id})
                temp.update(d)

            elif line.startswith(" nameif"):
                interface_name = line.split(" ")[2]
                d.update({"interface_name": interface_name})
                temp.update(d)

            elif line.startswith(" ip address"):
                ip_address = line.split(" ")[3] + " " + line.split(" ")[4]
                d.update({"ip": ip_address})
                temp.update(d)

            elif line.startswith(" channel-group"):
                channel_group = line.split(" ")[2]
                d.update({"channel-group": channel_group})
                temp.update(d)

            elif line.startswith(" shutdown"):
                status = "down"
                d.update({"status": status})
                temp.update(d)

            elif line.startswith("!"):
                for i in range(0, 5):

                    if "vlan" not in temp:
                        vlan_id = None
                        temp.update({"vlan": vlan_id})

                    elif "interface_name" not in temp:
                        interface_name = "no_interface_name"
                        temp.update({"interface_name": interface_name})

                    elif "ip" not in temp:
                        ip_address = None
                        temp.update({"ip": ip_address})

                    elif "channel-group" not in temp:
                        channel_group = None
                        temp.update({"channel-group": channel_group})

                    elif "status" not in temp:
                        status = "up"
                        temp.update({"status": status})

                output_txt_list.append(temp)
                del temp
                temp = {}

        return output_txt_list
    except Exception as e:
        print("test_interface failed!")
        print(str(e))

def output_excel(interfaces,ordered_list,output):

    wb=Workbook(output)
    ws=wb.add_worksheet() # Or leave it blank. The default name is "Sheet 1"
    
    first_row=0
    for header in ordered_list:
        col=ordered_list.index(header) # We are keeping order.
        ws.write(first_row,col,header) # We have written first row which is the header of worksheet also.
    
    row=1
    for interface in interfaces:
        for _key,_value in interface.items():
            col=ordered_list.index(_key)
            ws.write(row,col,_value)
        row+=1 #enter the next row
    wb.close()


def split_member_interface(text):
    text = text.replace('"',"")
    text = text.replace('[',"")
    text = text.replace(']',"")
    text = text.replace(' ',"")
    
    members = text.split(",")
    members1 = [value for value in members if value != ""]
    output = []
    for x in members1:
        output.append({"interface-name": x})
    return output



def config_with_netmiko(fgtip,interface_name, interface_type, vdom, interface_role, interface_allow_access, interface_vlanid,ip,interface,status,session):
    if ip==None:
        ip = "0.0.0.0/0"
    if interface_vlanid ==None:
        interface_vlanid = "0"

    config_set = ['end','config vdom', 'edit '+vdom,'config system interface', 'edit '+interface_name, 'set ip '+ip,'set allowaccess '+interface_allow_access,'set role '+ interface_role, 'set type '+interface_type,'next','end']
    #print (config_set)
    session.send_config_set(config_set)


    
def create_interface(fgtip,interface_name, interface_type, vdom, interface_role, interface_allow_access, member, interface_vlanid,ip,interface,status):
    interface_link = '/system/interface'
    #interface_url = "https://" + fgtip + "/api/v2/cmdb" + interface_link + "?access_token=" + token + "&vdom=" + vdom 
    interface_url = "https://" + fgtip + "/api/v2/cmdb" + interface_link + "?access_token=" + token  + "&vdom=" + vdom  
    #print(interface_name)
    #print (interface_url)

    if member!=None:
        new_member = split_member_interface(member)
        
    #print(interface_name)
    

    
    
    '''
    if interface_type == 'physical':
        config_with_netmiko(fgtip,interface_name, interface_type, vdom, interface_role, interface_allow_access, member, interface_vlanid,ip,interface,status)
    '''      
    if interface_type == 'physical':
        interface_data = {
                "name": interface_name,
                "type": interface_type,
                "vdom": vdom,
                "role": interface_role,
                "allowaccess": interface_allow_access,
                "vlanid": interface_vlanid,
                "ip": ip,
                "interface": interface,
                "status": status
            }    
    elif interface_type == 'vlan':
        interface_data = {
                "name": interface_name,
                "type": interface_type,
                "vdom": vdom,
                "role": interface_role,
                "allowaccess": interface_allow_access,
                "vlanid": interface_vlanid,
                "ip": ip,
                "interface": interface,
                "status": status
            }

        
    elif interface_type == 'aggregate':
        interface_data = {
                "name": interface_name,
                "type": interface_type,
                "vdom": vdom,
                "role": interface_role,
                "allowaccess": interface_allow_access,
                "member": new_member,
                "vlanid": interface_vlanid,
                "ip": ip,
                "interface": interface,
                "status": status
            }

    #print(interface_data)
    interface_body = json.dumps(interface_data)
    #print(interface_data)
    create_result = requests.post(url=interface_url, data=interface_body, verify=False)
    return create_result.status_code

def test_accessgroup(in_text):
    try:
        input_txt_list = in_text.split("\n")
        output_txt_list = []
        temp = {}
        for line in input_txt_list:
            d = {}

            if line.startswith("access-group"):
                
                access_group_name = line.split(" ")[1]
                if access_group_name!= 'global_access':
                    name_if = line.split(" ")[4]
                    #print (access_group_name)
                    #print (name_if)
                    
                    d.update({"access-group-name": access_group_name})
                    d.update({"interface_name_asa": name_if})
                    #print(d)
            temp.update(d)
            output_txt_list.append(temp)
            del temp
            temp = {}
        #print (temp)
        output_txt_list = [value for value in output_txt_list if value != {}]
        return output_txt_list
    except Exception as e:
        print("test_interface failed!")
        print(str(e))


def merge_lists(list1,list2):
    for i in list1:
        for j in list2:
            if i['interface_name_asa'] == j['interface_name_asa']:
                i.update(j)
    return list1

def create_sublist(list1):
    
    list2=[]
    d={}
    for i in list1:
        if find_type_interface(i['interface_id']) == 'portchannel_interface_type':
            
            d.update({'id':i['interface_id']})
            d.update({'name':i['interface_name']})
            list2.append(d)
            del d
            d = {}
    return list2
    
def replace_list(list1,list2):
    for i in list1:
        for j in list2:
            if i["interface"] != None:
                if i["interface"] in j["id"]:
                    i["interface"] = j['name']
    
    return list1
    
    

if __name__ == '__main__':
    
    
    asa_ip = input("Enter your ASA ip : ")
    user = input("Enter your username : ")
    password = getpass.getpass()
    excel_name = input("Enter name of excel output of interface : ")
    fgtip = input("Enter your Fortigate ip : ")
    user_forti= input("Enter your Fortigate username : ")
    password_forti = getpass.getpass()
    vdom = input("Enter vdom name of Fortigate: ")
    token = input("Enter token of Fortigate : ")

    session_asa = ConnectHandler(device_type='cisco_ios',host=asa_ip, username=user, password=password , secret=password)    
    session_asa.enable()
    session_asa.send_command("terminal pager 0")
    InterfaceASA = session_asa.send_command("show running-config interface")
    dict_list = test_interface(InterfaceASA)
    edited_dict_list = edit_dict_list(dict_list)


    AccessGroupASA = session_asa.send_command("show running-config | include access-group")
    dict_list_access_group = test_accessgroup(AccessGroupASA)
    
    edited_dict_list = merge_lists(edited_dict_list,dict_list_access_group)
    portchannel_list = create_sublist(edited_dict_list)

    edited_dict_list = replace_list(edited_dict_list,portchannel_list)
    output_excel(edited_dict_list,["interface_id","channel-group","vlan", "interface_name","interface_allow_access", "ip","status","interface","role","type","interface_name_asa","access-group-name"],excel_name)
    
    wb = load_workbook(excel_name, data_only=True)
    sheet = wb["Sheet1"]
    
    
    print(Fore.CYAN +"I try to connect fortigate, Please wait ..." + Style.RESET_ALL)
    session = ConnectHandler(device_type='fortinet',host=fgtip, username=user_forti, password=password_forti )
    
    
    
    for row in sheet.iter_rows(min_row=2,min_col=1,max_col=20,values_only=True):

        member = row[1]
        interface_vlanid = row[2]
        interface_name = row[3]
        interface_allow_access = row[4]
        ip = row[5]
        status = row[6]
        interface = row[7]
        interface_role = row[8]
        interface_type = row[9]
        
        result_code = create_interface(fgtip,interface_name, interface_type, vdom, interface_role, interface_allow_access, member, interface_vlanid,ip,interface,status)
        
        if result_code!=200:
            if interface_type == 'physical' and interface_name !=None:
                print (Fore.RED +"I can't create interface %s with Restconf"%interface_name + Style.RESET_ALL)
                config_with_netmiko(fgtip,interface_name, interface_type, vdom, interface_role, interface_allow_access, interface_vlanid,ip,interface,status,session)
                print (Fore.GREEN +"I create interface %s with Netmiko"%interface_name + Style.RESET_ALL)
            elif interface_type == 'physical' and interface_name ==None:
                pass
            else:
                print (Fore.RED + "Create Interface %s" %interface_name ," on host %s" %fgtip,"is %s"%result_code+ Style.RESET_ALL)
        else:
            print (Fore.GREEN + "Create Interface %s" %interface_name ," on host %s" %fgtip,"is successfully"+ Style.RESET_ALL)
    
