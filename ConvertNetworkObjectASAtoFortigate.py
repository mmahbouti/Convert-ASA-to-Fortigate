import getpass
import sys
import colorama
from colorama import init
from netmiko import ConnectHandler
init()
from colorama import Fore,Back,Style
from openpyxl import load_workbook
import re
import requests
import urllib3
import json
import xlsxwriter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def extract_data2(text_network): 
    try:
        output = ""
        input_txt_list = text_network.split("\n")
        
        for line in input_txt_list:

            if line.startswith("object network "):
                line_list = line.split(" ")
                
                if " host " in line:
                    output = output + line_list[2] + " host " + line_list[4] +"\n"
                elif " fqdn " in line:
                    if line_list[4]=="v4" or line_list[4]=="v6":
                        output = output + line_list[2] + " fqdn " + line_list[5] +"\n"
                    else:
                        output = output + line_list[2] + " fqdn " + line_list[4] +"\n"
                        
                elif " subnet " in line:
                    output = output + line_list[2] + " subnet " + line_list[4]+ "-" + line_list[5] +"\n"
                elif " range " in line:
                    output = output + line_list[2] + " range " + line_list[4]+ "-" + line_list[5] +"\n"

        return output

    except Exception as e:
        print("extract_data failed!")
        print(str(e))
		
def create_output_excel(text,header_list,output_name):
    workbook = xlsxwriter.Workbook(output_name)
    worksheet = workbook.add_worksheet()
    
    #write header on row=0
    for col_num, data in enumerate(header_list):
        worksheet.write(0, col_num, data)
        
    t1 = text.split("\n")
    l=[]
    for x in t1:
        
        y= x.split(" ")
        l.append(y)
        
    for row_num, row_data in enumerate(l):
        for col_num, col_data in enumerate(row_data):
            worksheet.write(row_num+1, col_num, col_data)
    
    workbook.close()
    return output_name

    
def create_address_object(name,type_address,address,fgtip,token,vdom):

    address_object_link = '/firewall/address'
    address_object_url = "https://" + fgtip + "/api/v2/cmdb" + address_object_link  + "?access_token=" + token + vdom 
    
    if type_address== 'host':
        address = address + " 255.255.255.255"
        type_address = "ipmask"
        
        address_data = {
        "name": name,
        "type": type_address,
        "subnet": address       
        }
        
    elif type_address== 'fqdn':
        type_address = "fqdn"
        
        address_data = {
        "name": name,
        "type": type_address,
        "fqdn": address
        }
        
    elif type_address== 'range':
        address_new = address.split("-")
        start_ip = address_new[0]
        end_ip = address_new[1]
        type_address = "iprange"
        
        address_data = {
        "name": name,
        "type": type_address,
        "start-ip": start_ip,
        "end-ip": end_ip
        }

    
    elif type_address== 'subnet':
        address_new = address.split("-")
        net_id = address_new[0]
        subnet = address_new[1]
        address = net_id + " " + subnet
        type_address = "ipmask"
        
        address_data = {
        "name": name,
        "type": type_address,
        "subnet": address       
        }


    address_data_body = json.dumps(address_data)
    create_result = requests.post(url=address_object_url, data=address_data_body, verify=False)

    return create_result.status_code

def create_text(text1,text2):
    output=""
    output= output + text1 + " " + text2 + "\n"
    return output

def processString(txt):

    
    dictionary = {'(': '-', ')':'-', '#': '-'}
    transTable = txt.maketrans(dictionary)
    txt = txt.translate(transTable)
    
    if txt[-1] ==  '-':
        txt = txt[:-1]
    elif txt[0] ==  '-':
        txt = txt[1:]
    
    return txt
   
if __name__ == "__main__":

    asa_ip = input("Enter your ASA ip : ")
    user = input("Enter your username : ")
    password = getpass.getpass()
    file_name_excel_asa = input("Enter name of excel output of ASA network Object : ")

    fgtip = input("Enter your Fortigate ip : ")
    vdom_name = input("Enter vdom name of Fortigate: ")
    token = input("Enter token of Fortigate : ")
    file_name_excel_fortigate = input("Enter name of excel output of Fortigate network Object : ")
    file_name_excel_equalizer = input("Enter name of excel output of equalizer network Object : ")

    vdom = "&vdom=" + vdom_name

    session = ConnectHandler(device_type='cisco_ios',host=asa_ip, username=user, password=password , secret=password)    
    session.enable()
    session.send_command("terminal pager 0")
    NetworkObjectASA = session.send_command("show running-config object network in-line")
    
    NetworkObjectASA = extract_data2(NetworkObjectASA)

    create_output_excel(NetworkObjectASA,['name','type','address'],file_name_excel_asa)
    
    wb = load_workbook(file_name_excel_asa, data_only=True)
    sheet = wb["Sheet1"]
    ResultCreateNetworkObject=""
    EqualizerNetworkObject = ""
    for row in sheet.iter_rows(min_row=2,min_col=1,max_col=20,values_only=True):
        
        name = row[0]
        type_address = row[1]
        address = row[2]

        result_code = create_address_object(name,type_address,address,fgtip,token,vdom)
        
        
        
        if result_code != 200:
            
            new_name = processString(name)
            result_code = create_address_object(new_name,type_address,address,fgtip,token,vdom)
            if result_code != 200:
                print (Fore.RED + "Create network object %s" %new_name ," with address %s"%address ," on host %s" %fgtip,"is %s"%result_code+ Style.RESET_ALL)
            else:
                EqualizerNetworkObject = EqualizerNetworkObject + name + " " + new_name + "\n"
                print (Fore.GREEN + "Create network object %s" %new_name ," with address %s"%address ," on host %s" %fgtip,"is successfully"+ Style.RESET_ALL)

        else:
            EqualizerNetworkObject = EqualizerNetworkObject + name + " " + name + "\n"
            print (Fore.GREEN + "Create network object %s" %name ," with address %s"%address ," on host %s" %fgtip,"is successfully"+ Style.RESET_ALL)

        ResultCreateNetworkObject =  ResultCreateNetworkObject + name + " " + type_address + " " + address + " " + str(result_code) + "\n"

    create_output_excel(ResultCreateNetworkObject,['name','type','address','result code'],file_name_excel_fortigate)
    create_output_excel(EqualizerNetworkObject,['name asa','name fortigate'],file_name_excel_equalizer)
